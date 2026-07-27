"""Bulk-load the Genpact F&A workbook into the ``genpact_*`` analytics tables.

Reads every structured sheet of
``data/Genpact_FA_Synthetic_Employee_Dataset.xlsx`` and inserts the rows
into its dedicated table (see :mod:`scripts._genpact_spec`), scoped to the
``genpact`` tenant. The four yearly benchmark sheets are unioned into
``genpact_benchmark``. Purely narrative reference sheets (Supply_Sources,
Summary_Dashboard, Comp_Benchmarks, Audit_Notes) are documentation and are
intentionally skipped.

Idempotent: clears the tenant's rows in every target table, then reloads.

Run::

    uv run python -m scripts.seed_genpact_master_data
"""

from __future__ import annotations

import asyncio
import time

import openpyxl
from app.core.database import AsyncSessionLocal
from app.models.genpact_master_data import GENPACT_TABLES
from sqlalchemy import delete

from scripts._genpact_common import EMPLOYEE_WORKBOOK, get_or_create_tenant
from scripts._genpact_spec import (
    BENCHMARK_SHEETS,
    BENCHMARK_TABLE,
    TABLES,
    coerce,
)

_BATCH = 5000


def _rows_for_sheet(ws, header_row: int, columns) -> list[dict]:
    """Coerce every data row of ``ws`` into a list of column dicts."""
    out: list[dict] = []
    for r_idx, raw in enumerate(ws.iter_rows(values_only=True)):
        if r_idx <= header_row:
            continue
        # Skip fully-blank rows.
        if raw is None or all(c is None or (isinstance(c, str) and c.strip() == "") for c in raw):
            continue
        row: dict = {}
        for col_idx, (attr, kind, _lp) in enumerate(columns):
            value = raw[col_idx] if col_idx < len(raw) else None
            row[attr] = coerce(kind, value)
        out.append(row)
    return out


async def _load_table(db, *, tenant_id, table, rows: list[dict]) -> int:
    await db.execute(delete(table).where(table.c.tenant_id == tenant_id))
    for i in range(0, len(rows), _BATCH):
        batch = [{"tenant_id": tenant_id, **row} for row in rows[i : i + _BATCH]]
        await db.execute(table.insert(), batch)
    return len(rows)


async def main() -> None:
    t0 = time.perf_counter()
    print(f"Opening workbook {EMPLOYEE_WORKBOOK.name} ...")
    wb = openpyxl.load_workbook(EMPLOYEE_WORKBOOK, read_only=True, data_only=True)

    async with AsyncSessionLocal() as db:
        tenant = await get_or_create_tenant(db)
        print(f"Tenant {tenant.code} ({tenant.id})")

        # Per-sheet structured tables.
        for sheet_name, (table_name, header_row, columns) in TABLES.items():
            if sheet_name not in wb.sheetnames:
                print(f"  ! sheet {sheet_name!r} missing — skipped")
                continue
            ws = wb[sheet_name]
            rows = _rows_for_sheet(ws, header_row, columns)
            n = await _load_table(
                db,
                tenant_id=tenant.id,
                table=GENPACT_TABLES[table_name],
                rows=rows,
            )
            print(f"  {table_name:<28} {n:>7,} rows")

        # Unioned benchmark table (4 yearly sheets -> one table).
        bench_table_name, bench_header, bench_cols = BENCHMARK_TABLE
        bench_rows: list[dict] = []
        for sheet_name in BENCHMARK_SHEETS:
            if sheet_name not in wb.sheetnames:
                print(f"  ! sheet {sheet_name!r} missing — skipped")
                continue
            bench_rows.extend(_rows_for_sheet(wb[sheet_name], bench_header, bench_cols))
        n = await _load_table(
            db,
            tenant_id=tenant.id,
            table=GENPACT_TABLES[bench_table_name],
            rows=bench_rows,
        )
        print(f"  {bench_table_name:<28} {n:>7,} rows (4 sheets unioned)")

        await db.commit()

    wb.close()
    print(f"Done in {time.perf_counter() - t0:.1f}s")


if __name__ == "__main__":
    asyncio.run(main())
